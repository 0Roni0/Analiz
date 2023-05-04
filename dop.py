import openpyxl
import numpy as np
from matplotlib import pyplot as plt
import os

wb = openpyxl.load_workbook(r'E:\analiz\analiz.xlsx')

def Myn(): 
        sheet = wb["Муниципальный этап"]
        print("Муниципальный этап")
        a = [sheet.cell(row=i,column=5).value for i in range(3,8)]
        b = [sheet.cell(row=i,column=9).value for i in range(3,8)]
        x = (np.divide(b,a) * 100)
        print("Физика =",x)
        m = [F'Год{i+1}' for i in range(2018, 2023)]
        plt.bar(m, a, alpha=0.3, label='Количество участников')
        plt.bar(m, b, alpha=0.6, label='Количество победителей и призеров')
        plt.legend(loc='upper right')
        plt.show()

        a = [sheet.cell(row=i,column=5).value for i in range(12,17)]
        b = [sheet.cell(row=i,column=9).value for i in range(12,17)]
        x = (np.divide(b,a) * 100)
        print("Информатика =", x)

        m = [F'Год{i+1}' for i in range(2018, 2023)]
        plt.bar(m, a, alpha=0.3, label='Количество участников')
        plt.bar(m, b, alpha=0.6, label='Количество победителей и призеров')
        plt.legend(loc='upper right')
        plt.show()

        a = [sheet.cell(row=i,column=5).value for i in range(21,26)]
        b = [sheet.cell(row=i,column=9).value for i in range(21,26)]
        x = (np.divide(b,a) * 100)
        print("Математика =",x)

        m = [F'Год{i+1}' for i in range(2018, 2023)]
        plt.bar(m, a, alpha=0.3, label='Количество участников')
        plt.bar(m, b, alpha=0.6, label='Количество победителей и призеров')
        plt.legend(loc='upper right')
        plt.show()

def Stav():
    sheet = wb["Ставропльский край"]
    print("Ставропльский край")
    a = [sheet.cell(row=i,column=2).value for i in range(3,7)]
    b = [sheet.cell(row=i,column=3).value for i in range(3,7)]
    x = (np.divide(b,a) * 100)
    print("Физика =",x)
    a = [sheet.cell(row=i,column=2).value for i in range(10,14)]
    b = [sheet.cell(row=i,column=3).value for i in range(10,14)]
    x = (np.divide(b,a) * 100)
    print("Информатика =",x)
    a = [sheet.cell(row=i,column=2).value for i in range(17,21)]
    b = [sheet.cell(row=i,column=3).value for i in range(17,21)]
    x = (np.divide(b,a) * 100)
    print("Математика =",x)

    m = [F'Год{i+1}' for i in range(2019, 2023)]
    plt.bar(m, a, alpha=0.3, label='Количество участников')
    plt.bar(m, b, alpha=0.6, label='Количество победителей и призеров')
    plt.legend(loc='upper right')
    plt.show()

def Rost():
    sheet = wb["Ростовская область"]
    print("Ростовская область")
    a = [sheet.cell(row=i,column=2).value for i in range(3,7)]
    b = [sheet.cell(row=i,column=3).value for i in range(3,7)]
    x = (np.divide(b,a) * 100)
    print("Физика =",x)
    a = [sheet.cell(row=i,column=2).value for i in range(10,14)]
    b = [sheet.cell(row=i,column=3).value for i in range(10,14)]
    x = (np.divide(b,a) * 100)
    print("Информатика =",x)
    a = [sheet.cell(row=i,column=2).value for i in range(17,21)]
    b = [sheet.cell(row=i,column=3).value for i in range(17,21)]
    x = (np.divide(b,a) * 100)
    print("Математика =",x)

    m = [F'Год{i+1}' for i in range(2019, 2023)]
    plt.bar(m, a, alpha=0.3, label='Количество участников')
    plt.bar(m, b, alpha=0.6, label='Количество победителей и призеров')
    plt.legend(loc='upper right')
    plt.show()

def Kras():
    sheet = wb["Краснодарский край"]
    print("Краснодарский край")
    a = [sheet.cell(row=i,column=2).value for i in range(3,7)]
    b = [sheet.cell(row=i,column=3).value for i in range(3,7)]
    x = (np.divide(b,a) * 100)
    print("Информатика =",x)
    a = [sheet.cell(row=i,column=2).value for i in range(10,14)]
    b = [sheet.cell(row=i,column=3).value for i in range(10,14)]
    x = (np.divide(b,a) * 100)
    print("Физика =",x)
    a = [sheet.cell(row=i,column=2).value for i in range(17,21)]
    b = [sheet.cell(row=i,column=3).value for i in range(17,21)]
    x = (np.divide(b,a) * 100)
    print("Математика =",x)

    m = [F'Год{i+1}' for i in range(2019, 2023)]
    plt.bar(m, a, alpha=0.3, label='Количество участников')
    plt.bar(m, b, alpha=0.6, label='Количество победителей и призеров')
    plt.legend(loc='upper right')
    plt.show()

def Dages():
    sheet = wb["Республика Дагестан"]
    print("Республика Дагестан")
    a = [sheet.cell(row=i,column=2).value for i in range(3,7)]
    b = [sheet.cell(row=i,column=3).value for i in range(3,7)]
    x = (np.divide(b,a) * 100)
    print("Информатика =",x)
    a = [sheet.cell(row=i,column=2).value for i in range(10,14)]
    b = [sheet.cell(row=i,column=3).value for i in range(10,14)]
    x = (np.divide(b,a) * 100)
    print("Физика =",x)
    a = [sheet.cell(row=i,column=2).value for i in range(17,21)]
    b = [sheet.cell(row=i,column=3).value for i in range(17,21)]
    x = (np.divide(b,a) * 100)
    print("Математика =",x)

    m = [F'Год{i+1}' for i in range(2019, 2023)]
    plt.bar(m, a, alpha=0.3, label='Количество участников')
    plt.bar(m, b, alpha=0.6, label='Количество победителей и призеров')
    plt.legend(loc='upper right')
    plt.show()

def Sever():
    sheet = wb["Республика Дагестан"]
    print("Республика Дагестан")
    a = [sheet.cell(row=i,column=2).value for i in range(3,7)]
    b = [sheet.cell(row=i,column=3).value for i in range(3,7)]
    x = (np.divide(b,a) * 100)
    print("Информатика =",x)
    a = [sheet.cell(row=i,column=2).value for i in range(10,14)]
    b = [sheet.cell(row=i,column=3).value for i in range(10,14)]
    x = (np.divide(b,a) * 100)
    print("Физика =",x)
    a = [sheet.cell(row=i,column=2).value for i in range(17,21)]
    b = [sheet.cell(row=i,column=3).value for i in range(17,21)]
    x = (np.divide(b,a) * 100)
    print("Математика =",x)

    m = [F'Год{i+1}' for i in range(2019, 2023)]
    plt.bar(m, a, alpha=0.3, label='Количество участников')
    plt.bar(m, b, alpha=0.6, label='Количество победителей и призеров')
    plt.legend(loc='upper right')
    plt.show()

def Kara():
    sheet = wb["Республика Дагестан"]
    print("Республика Дагестан")
    a = [sheet.cell(row=i,column=2).value for i in range(3,7)]
    b = [sheet.cell(row=i,column=3).value for i in range(3,7)]
    x = (np.divide(b,a) * 100)
    print("Информатика =",x)
    a = [sheet.cell(row=i,column=2).value for i in range(10,14)]
    b = [sheet.cell(row=i,column=3).value for i in range(10,14)]
    x = (np.divide(b,a) * 100)
    print("Физика =",x)
    a = [sheet.cell(row=i,column=2).value for i in range(17,21)]
    b = [sheet.cell(row=i,column=3).value for i in range(17,21)]
    x = (np.divide(b,a) * 100)
    print("Математика =",x)

    m = [F'Год{i+1}' for i in range(2019, 2023)]
    plt.bar(m, a, alpha=0.3, label='Количество участников')
    plt.bar(m, b, alpha=0.6, label='Количество победителей и призеров')
    plt.legend(loc='upper right')
    plt.show()

def Kalm():
    sheet = wb["Республика Калмыкия"]
    print("Республика Калмыкия")
    a = [sheet.cell(row=i,column=2).value for i in range(3,7)]
    b = [sheet.cell(row=i,column=3).value for i in range(3,7)]
    x = (np.divide(b,a) * 100)
    print("Информатика =",x)
    a = [sheet.cell(row=i,column=2).value for i in range(10,14)]
    b = [sheet.cell(row=i,column=3).value for i in range(10,14)]
    x = (np.divide(b,a) * 100)
    print("Физика =",x)
    a = [sheet.cell(row=i,column=2).value for i in range(17,21)]
    b = [sheet.cell(row=i,column=3).value for i in range(17,21)]
    x = (np.divide(b,a) * 100)
    print("Математика =",x)

    m = [F'Год{i+1}' for i in range(2019, 2023)]
    plt.bar(m, a, alpha=0.3, label='Количество участников')
    plt.bar(m, b, alpha=0.6, label='Количество победителей и призеров')
    plt.legend(loc='upper right')
    plt.show()

def Cheche():
    sheet = wb["Чеченская республика"]
    print("Чеченская республика")
    a = [sheet.cell(row=i,column=2).value for i in range(3,7)]
    b = [sheet.cell(row=i,column=3).value for i in range(3,7)]
    x = (np.divide(b,a) * 100)
    print("Информатика =",x)
    a = [sheet.cell(row=i,column=2).value for i in range(10,14)]
    b = [sheet.cell(row=i,column=3).value for i in range(10,14)]
    x = (np.divide(b,a) * 100)
    print("Физика =",x)
    a = [sheet.cell(row=i,column=2).value for i in range(17,21)]
    b = [sheet.cell(row=i,column=3).value for i in range(17,21)]
    x = (np.divide(b,a) * 100)
    print("Математика =",x)

    m = [F'Год{i+1}' for i in range(2019, 2023)]
    plt.bar(m, a, alpha=0.3, label='Количество участников')
    plt.bar(m, b, alpha=0.6, label='Количество победителей и призеров')
    plt.legend(loc='upper right')
    plt.show()

def Analiz():

    vibr = input('''
        [1] Анализ 2019-2020год.
        [2] Анализ 2020-2021год.
        [3] Анализ 2021-2022год.
        [4] Анализ 2022-2023год.
        [0] Выход
        Введите значение:''')

    def an19(): 
        sheet = wb["Анализ"]
        print("Анализ 2019-2020")
        a = [sheet.cell(row=i,column=4).value for i in range(4,13)]
        b = [sheet.cell(row=i,column=5).value for i in range(4,13)]
        x = (np.divide(b,a) * 100)
        print("Информатика =",x)
        a = [sheet.cell(row=i,column=7).value for i in range(4,13)]
        b = [sheet.cell(row=i,column=8).value for i in range(4,13)]
        x = (np.divide(b,a) * 100)
        print("Физика =",x)
        a = [sheet.cell(row=i,column=10).value for i in range(4,13)]
        b = [sheet.cell(row=i,column=11).value for i in range(4,13)]
        x = (np.divide(b,a) * 100)
        print("Математика =",x)

        m = [F'Рег{i+1}' for i in range(9)]
        plt.bar(m, a, alpha=0.3, label='Количество участников')
        plt.bar(m, b, alpha=0.6, label='Количество победителей и призеров')
        plt.legend(loc='upper right')
        plt.show()

    def an20():
        sheet = wb["Анализ"]
        print("Анализ 2020-2021")
        a = [sheet.cell(row=i,column=4).value for i in range(18,27)]
        b = [sheet.cell(row=i,column=5).value for i in range(18,27)]
        x = (np.divide(b,a) * 100)
        print("Информатика =",x)
        a = [sheet.cell(row=i,column=7).value for i in range(18,27)]
        b = [sheet.cell(row=i,column=8).value for i in range(18,27)]
        x = (np.divide(b,a) * 100)
        print("Физика =",x)
        a = [sheet.cell(row=i,column=10).value for i in range(18,27)]
        b = [sheet.cell(row=i,column=11).value for i in range(18,27)]
        x = (np.divide(b,a) * 100)
        print("Математика =",x)    

        m = [F'Рег{i+1}' for i in range(9)]
        plt.bar(m, a, alpha=0.3, label='Количество участников')
        plt.bar(m, b, alpha=0.6, label='Количество победителей и призеров')
        plt.legend(loc='upper right')
        plt.show()

    def an21():
        sheet = wb["Анализ"]
        print("Анализ 2021-2022")
        a = [sheet.cell(row=i,column=4).value for i in range(32,41)]
        b = [sheet.cell(row=i,column=5).value for i in range(32,41)]
        x = (np.divide(b,a) * 100)
        print("Информатика =",x)
        a = [sheet.cell(row=i,column=7).value for i in range(32,41)]
        b = [sheet.cell(row=i,column=8).value for i in range(32,41)]
        x = (np.divide(b,a) * 100)
        print("Физика =",x)
        a = [sheet.cell(row=i,column=10).value for i in range(32,41)]
        b = [sheet.cell(row=i,column=11).value for i in range(32,41)]
        x = (np.divide(b,a) * 100)
        print("Математика =",x)   

        m = [F'Рег{i+1}' for i in range(9)]
        plt.bar(m, a, alpha=0.3, label='Количество участников')
        plt.bar(m, b, alpha=0.6, label='Количество победителей и призеров')
        plt.legend(loc='upper right')
        plt.show()

    def an22():
        sheet = wb["Анализ"]
        print("Анализ 2022-2023")
        a = [sheet.cell(row=i,column=4).value for i in range(46,55)]
        b = [sheet.cell(row=i,column=5).value for i in range(46,55)]
        x = (np.divide(b,a) * 100)
        print("Информатика =",x)
        a = [sheet.cell(row=i,column=7).value for i in range(46,55)]
        b = [sheet.cell(row=i,column=8).value for i in range(46,55)]
        x = (np.divide(b,a) * 100)
        print("Физика =",x)
        a = [sheet.cell(row=i,column=10).value for i in range(46,55)]
        b = [sheet.cell(row=i,column=11).value for i in range(46,55)]
        x = (np.divide(b,a) * 100)
        print("Математика =",x)

        m = [F'Рег{i+1}' for i in range(9)]
        plt.bar(m, a, alpha=0.3, label='Количество участников')
        plt.bar(m, b, alpha=0.6, label='Количество победителей и призеров')
        plt.legend(loc='upper right')
        plt.show()

    if vibr == "1":
        an19()
    elif vibr == "2":
        an20()
    elif vibr == "3":
        an21()
    elif vibr == "4":
        an22()
    elif vibr == "0":
        quit("Выход")
    else:
        print("Такого значения нет")
        clear = lambda: os.system('cls')
        clear()
        Analiz()